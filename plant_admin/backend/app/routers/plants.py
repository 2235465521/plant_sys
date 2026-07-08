import json
import secrets
import urllib.error
from io import BytesIO
from pathlib import Path
from typing import Annotated, Any, Literal, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill

import pymysql
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse

from app.auth_utils import get_current_user, require_admin
from app.config import get_settings
from app.database import get_db
from app.models import PlantAdminUser
from app.plant_image_fetch import (
    download_image_bytes,
    resolve_remote_image_url,
)
from app.schemas import (
    PlantCreate,
    PlantDeleteServerImageIn,
    PlantExportBatchIn,
    PlantListOut,
    PlantOut,
    PlantUpdate,
    TaxonBucket,
    TaxonomyDistinctOut,
)
from app.semantic_search import query_semantic_similarity

router = APIRouter(prefix="/plants", tags=["plants"])

_PLANT_UPLOAD_MAX_BYTES = 20 * 1024 * 1024
_UPLOAD_MIME_TO_EXT: dict[str, str] = {
    "image/jpeg": ".jpg",
    "image/jpg": ".jpg",
    "image/png": ".png",
    "image/webp": ".webp",
    "image/gif": ".gif",
}

_MAX_MEDIA_PER_PLANT = 80


def _deserialize_paths_from_db(raw: Any) -> Optional[list[str]]:
    if raw is None:
        return None
    if isinstance(raw, bytes):
        raw = raw.decode("utf-8", errors="replace")
    if isinstance(raw, str):
        t = raw.strip()
        if not t:
            return None
        try:
            parsed = json.loads(t)
        except json.JSONDecodeError:
            return None
        if not isinstance(parsed, list):
            return None
        out = [str(x).strip() for x in parsed if x is not None and str(x).strip()]
        return out or None
    if isinstance(raw, list):
        out = [str(x).strip() for x in raw if x is not None and str(x).strip()]
        return out or None
    return None


def _serialize_paths_for_db(paths: Optional[list[str]]) -> Any:
    if not paths:
        return None
    return json.dumps(list(paths), ensure_ascii=False)


def _strip_url_query(raw: str) -> str:
    return raw.split("?", 1)[0].strip()


def _basename_under_plant_media(plant_id: int, public_path: str) -> Optional[str]:
    """校验 path 必须为 /api/media/plants/{id}/文件名，返回文件名。"""
    p = _strip_url_query(public_path)
    prefix = f"/api/media/plants/{plant_id}/"
    if not p.startswith(prefix):
        return None
    basename = p[len(prefix) :]
    if not basename or "/" in basename or "\\" in basename or basename.startswith("."):
        return None
    if ".." in basename:
        return None
    return basename


def _row_as_plant_out_dict(row: dict) -> dict:
    d = dict(row)
    d["image_server_paths"] = _deserialize_paths_from_db(d.get("image_server_paths"))
    return d


def plant_row_to_out(
    row: dict,
    aliases: list[dict] = None,
    habitats: list[str] = None,
    rankings: list[dict] = None,
    regions: list[dict] = None
) -> PlantOut:
    d = _row_as_plant_out_dict(row)
    if aliases is not None:
        d["aliases"] = aliases
    if habitats is not None:
        d["habitats"] = habitats
    if rankings is not None:
        d["rankings"] = rankings
    if regions is not None:
        d["regions"] = regions
    return PlantOut.model_validate(d)


def _sql_column_value(column: str, value: Any) -> Any:
    if column == "image_server_paths":
        return _serialize_paths_for_db(value if isinstance(value, list) else None)
    return value


_PLANT_COLS = (
    "division",
    "subclass",
    "taxonomic_order",
    "family",
    "genus",
    "vernacular_name",
    "alternative_names_zh",
    "scientific_name",
    "taxonomic_provenance",
    "synonyms",
    "morphology_text",
    "medicinal_shape",
    "distribution_china",
    "distribution_abroad",
    "habitat",
    "medicinal_part",
    "is_medicinal_food_homologous",
    "image_url",
    "image_server_paths",
    "harvest_months",
    "food_therapy_months",
    "harvest_months_desc",
    "food_therapy_months_desc",
)


def _sync_plant_auto_increment(conn: pymysql.connections.Connection) -> None:
    """与当前 MAX(id) 对齐，避免 AI 计数器超前导致中间号像被锁住。"""
    with conn.cursor() as cur:
        cur.execute("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM plant_classification_import")
        next_id = int((cur.fetchone() or {"next_id": 1})["next_id"])
        cur.execute(
            "ALTER TABLE plant_classification_import AUTO_INCREMENT = %s",
            (next_id,),
        )
    conn.commit()


def _enrich_rows_with_nested_features(rows: list[dict], conn: pymysql.connections.Connection) -> None:
    if not rows:
        return
    
    plant_ids = [r["id"] for r in rows]
    ph = ",".join(["%s"] * len(plant_ids))
    
    # 1. Aliases
    aliases_map = {}
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT plant_id, alias_type, alias_name, origin_desc FROM plant_aliases WHERE plant_id IN ({ph})",
            plant_ids
        )
        for al in cur.fetchall() or []:
            pid = al["plant_id"]
            desc_part = f" (由来: {al['origin_desc']})" if al.get("origin_desc") else ""
            val = f"[{al['alias_type']}] {al['alias_name']}{desc_part}"
            aliases_map.setdefault(pid, []).append(val)
            
    # 2. Habitats
    habitats_map = {}
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT plant_id, habitat_type FROM plant_habitats WHERE plant_id IN ({ph})",
            plant_ids
        )
        for h in cur.fetchall() or []:
            pid = h["plant_id"]
            habitats_map.setdefault(pid, []).append(h["habitat_type"])
            
    # 3. Rankings
    rankings_map = {}
    ranking_type_cn = {
        "sweetest": "最甜植物",
        "bitterest": "最苦植物",
        "rarity": "珍稀物种",
        "growth_cycle": "特殊生长周期"
    }
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT plant_id, ranking_type, ranking_value, description FROM plant_rankings WHERE plant_id IN ({ph})",
            plant_ids
        )
        for r in cur.fetchall() or []:
            pid = r["plant_id"]
            type_label = ranking_type_cn.get(r["ranking_type"], r["ranking_type"])
            val_part = f" (指标: {r['ranking_value']})" if r.get("ranking_value") else ""
            desc_part = f" 理由: {r['description']}" if r.get("description") else ""
            detail = f"{type_label}{val_part}{desc_part}"
            rankings_map.setdefault(pid, []).append(detail)
            
    # 4. Regions
    regions_map = {}
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT plant_id, region_name, combo_name FROM plant_regions WHERE plant_id IN ({ph})",
            plant_ids
        )
        for reg in cur.fetchall() or []:
            pid = reg["plant_id"]
            combo_part = f" ({reg['combo_name']})" if reg.get("combo_name") else ""
            val = f"{reg['region_name']}{combo_part}"
            regions_map.setdefault(pid, []).append(val)
            
    # Write back to each row
    for r in rows:
        pid = r["id"]
        r["aliases_str"] = "; ".join(aliases_map.get(pid, []))
        r["habitats_str"] = ", ".join(habitats_map.get(pid, []))
        r["rankings_str"] = "; ".join(rankings_map.get(pid, []))
        r["regions_str"] = "; ".join(regions_map.get(pid, []))


def plant_to_dict(p: dict) -> dict:
    return {
        "id": p["id"],
        "division": p.get("division"),
        "subclass": p.get("subclass"),
        "taxonomic_order": p.get("taxonomic_order"),
        "family": p.get("family"),
        "genus": p.get("genus"),
        "vernacular_name": p.get("vernacular_name"),
        "alternative_names_zh": p.get("alternative_names_zh"),
        "scientific_name": p.get("scientific_name"),
        "taxonomic_provenance": p.get("taxonomic_provenance"),
        "synonyms": p.get("synonyms"),
        "morphology_text": p.get("morphology_text"),
        "medicinal_shape": p.get("medicinal_shape"),
        "distribution_china": p.get("distribution_china"),
        "distribution_abroad": p.get("distribution_abroad"),
        "habitat": p.get("habitat"),
        "medicinal_part": p.get("medicinal_part"),
        "is_medicinal_food_homologous": p.get("is_medicinal_food_homologous"),
        "image_url": p.get("image_url"),
        "image_server_paths": _deserialize_paths_from_db(p.get("image_server_paths")),
        "harvest_months": p.get("harvest_months"),
        "food_therapy_months": p.get("food_therapy_months"),
        "harvest_months_desc": p.get("harvest_months_desc"),
        "food_therapy_months_desc": p.get("food_therapy_months_desc"),
        "aliases_str": p.get("aliases_str"),
        "habitats_str": p.get("habitats_str"),
        "rankings_str": p.get("rankings_str"),
        "regions_str": p.get("regions_str"),
    }


# 文本导出：字段名为中文，顺序固定；不含 JSON 括号。
_PLANT_EXPORT_EN_TO_CN: dict[str, str] = {
    "id": "编号",
    "division": "门",
    "subclass": "亚纲",
    "taxonomic_order": "目",
    "family": "科",
    "genus": "属",
    "vernacular_name": "中文名",
    "alternative_names_zh": "中文别名",
    "aliases_str": "分类别称系统",
    "scientific_name": "拉丁学名",
    "taxonomic_provenance": "分类来源或文献",
    "synonyms": "学名异名",
    "morphology_text": "形态描述",
    "medicinal_shape": "药用性状",
    "medicinal_part": "入药部位",
    "distribution_china": "国内分布",
    "distribution_abroad": "国外分布",
    "habitat": "生境",
    "habitats_str": "场景生境分类",
    "is_medicinal_food_homologous": "药食同源",
    "image_url": "参考图片网址",
    "image_server_paths": "本站存储图片路径列表",
    "harvest_months": "最佳采收月",
    "harvest_months_desc": "最佳采收说明",
    "food_therapy_months": "适合食疗入药月",
    "food_therapy_months_desc": "食疗入药说明",
    "rankings_str": "特色榜单荣誉",
    "regions_str": "道地中药材地",
}

_PLANT_EXPORT_FIELD_ORDER_EN: tuple[str, ...] = (
    "id",
    "division",
    "subclass",
    "taxonomic_order",
    "family",
    "genus",
    "vernacular_name",
    "alternative_names_zh",
    "aliases_str",
    "scientific_name",
    "taxonomic_provenance",
    "synonyms",
    "morphology_text",
    "medicinal_shape",
    "medicinal_part",
    "distribution_china",
    "distribution_abroad",
    "habitat",
    "habitats_str",
    "is_medicinal_food_homologous",
    "image_url",
    "image_server_paths",
    "harvest_months",
    "harvest_months_desc",
    "food_therapy_months",
    "food_therapy_months_desc",
    "rankings_str",
    "regions_str",
)

def plant_to_cn_export_row(p: dict) -> dict:
    d = plant_to_dict(p)
    return {_PLANT_EXPORT_EN_TO_CN.get(k, k): v for k, v in d.items()}


def _field_value_to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, list):
        return "; ".join(str(x) for x in val if x is not None)
    return str(val)


def plant_export_plain_text_block(p: dict) -> str:
    """单株植物导出为纯文本块（字段名: 值，每行一个字段）。"""
    lines: list[str] = []
    for field_en in _PLANT_EXPORT_FIELD_ORDER_EN:
        field_cn = _PLANT_EXPORT_EN_TO_CN.get(field_en, field_en)
        val = plant_to_dict(p).get(field_en)
        lines.append(f"{field_cn}: {_field_value_to_str(val)}")
    return "\n".join(lines)


def plants_export_plain_text(rows: list[dict]) -> str:
    """多株植物导出为纯文本，各株之间用分隔线隔开。"""
    blocks = [plant_export_plain_text_block(r) for r in rows]
    return "\n\n" + ("-" * 40) + "\n\n".join(blocks)


def plants_export_xlsx_bytes(rows: list[dict]) -> bytes:
    """多株植物导出为 Excel (.xlsx) 文件，返回字节流。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "植物数据"

    # Header row
    headers = [_PLANT_EXPORT_EN_TO_CN.get(f, f) for f in _PLANT_EXPORT_FIELD_ORDER_EN]
    ws.append(headers)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for p in rows:
        d = plant_to_dict(p)
        row = [_field_value_to_str(d.get(f)) for f in _PLANT_EXPORT_FIELD_ORDER_EN]
        ws.append(row)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _norm_str_list(raw: Optional[list[str]]) -> list[str]:
    if not raw:
        return []
    out: list[str] = []
    for x in raw:
        if x is None:
            continue
        s = str(x).strip()
        if s:
            out.append(s)
    return out


def _in_clause(col: str, values: list[str]) -> tuple[Optional[str], list]:
    if not values:
        return None, []
    ph = ",".join(["%s"] * len(values))
    return f"{col} IN ({ph})", list(values)


def _build_where(
    q: Optional[str],
    division: list[str],
    subclass: list[str],
    taxonomic_order: list[str],
    family: list[str],
    genus: list[str],
    harvest_month: Optional[int] = None,
    food_therapy_month: Optional[int] = None,
) -> tuple[str, list]:
    conds: list[str] = []
    params: list = []
    if q:
        kw = f"%{q.strip()}%"
        conds.append(
            "(vernacular_name LIKE %s OR scientific_name LIKE %s)"
        )
        params.extend([kw, kw])
    if division:
        sql, pr = _in_clause("division", division)
        if sql:
            conds.append(sql)
            params.extend(pr)
    if subclass:
        sql, pr = _in_clause("subclass", subclass)
        if sql:
            conds.append(sql)
            params.extend(pr)
    if taxonomic_order:
        sql, pr = _in_clause("taxonomic_order", taxonomic_order)
        if sql:
            conds.append(sql)
            params.extend(pr)
    if family:
        sql, pr = _in_clause("family", family)
        if sql:
            conds.append(sql)
            params.extend(pr)
    if genus:
        sql, pr = _in_clause("genus", genus)
        if sql:
            conds.append(sql)
            params.extend(pr)
    if harvest_month is not None:
        conds.append("FIND_IN_SET(%s, harvest_months) > 0")
        params.append(str(harvest_month))
    if food_therapy_month is not None:
        conds.append("FIND_IN_SET(%s, food_therapy_months) > 0")
        params.append(str(food_therapy_month))
    if not conds:
        return "1 = 1", []
    return " AND ".join(conds), params


@router.get("/taxonomy/distinct", response_model=TaxonomyDistinctOut)
def taxonomy_distinct(
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(get_current_user),
    field: Literal["division", "subclass", "taxonomic_order", "family", "genus"] = Query(
        ...,
        description="要获取去重取值的列（树形钻取：传入祖先字段可只返回其子级）",
    ),
    division: Optional[str] = Query(None, description="按门筛选子级"),
    subclass: Optional[str] = Query(None, description="按亚纲筛选子级"),
    torder: Optional[str] = Query(None, description="按目筛选子级"),
    family: Optional[str] = Query(None, description="按科筛选子级"),
):
    col = field
    conditions: list[str] = [f"({col} IS NOT NULL AND TRIM({col}) <> '')"]
    args: list[str] = []
    if division:
        conditions.append("TRIM(division) = %s")
        args.append(division.strip())
    if subclass:
        conditions.append("TRIM(subclass) = %s")
        args.append(subclass.strip())
    if torder:
        conditions.append("TRIM(taxonomic_order) = %s")
        args.append(torder.strip())
    if family:
        conditions.append("TRIM(family) = %s")
        args.append(family.strip())
    where = " AND ".join(conditions)
    sql = f"""
        SELECT TRIM({col}) AS v, COUNT(*) AS cnt
        FROM plant_classification_import
        WHERE {where}
        GROUP BY TRIM({col})
        ORDER BY v
        LIMIT 20000
    """
    with conn.cursor() as cur:
        if args:
            cur.execute(sql, args)
        else:
            cur.execute(sql)
        rows = cur.fetchall() or []
    items = [
        TaxonBucket(value=r["v"], count=int(r["cnt"]))
        for r in rows
        if r.get("v")
    ]
    return TaxonomyDistinctOut(field=field, items=items)


@router.get("/taxonomy/resolve-path")
def resolve_taxonomy_path(
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(get_current_user),
    level: Literal["division", "subclass", "taxonomic_order", "family", "genus"] = Query(...),
    value: str = Query(..., min_length=1),
):
    """给定某层级的一个分类名称，返回其完整溯源路径（取库中首条匹配记录的祖先字段）。"""
    col = level
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT division, subclass, taxonomic_order, family, genus "
            f"FROM plant_classification_import "
            f"WHERE TRIM({col}) = %s AND division IS NOT NULL LIMIT 1",
            (value.strip(),),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, f"未找到 {level}={value}")
    return {
        "division": (row.get("division") or "").strip() or None,
        "subclass": (row.get("subclass") or "").strip() or None,
        "taxonomic_order": (row.get("taxonomic_order") or "").strip() or None,
        "family": (row.get("family") or "").strip() or None,
        "genus": (row.get("genus") or "").strip() or None,
    }


@router.get("", response_model=PlantListOut)
def list_plants(
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(get_current_user),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    q: Optional[str] = Query(None, description="搜索中文名、拉丁名、属"),
    division: list[str] = Query(default=[], description="门，可多选"),
    subclass: list[str] = Query(default=[], description="亚纲，可多选"),
    taxonomic_order: list[str] = Query(default=[], alias="torder", description="目，可多选"),
    family: list[str] = Query(default=[], description="科，可多选"),
    genus: list[str] = Query(default=[], description="属，可多选"),
    harvest_month: Optional[int] = Query(None, description="最佳采收月筛选"),
    food_therapy_month: Optional[int] = Query(None, description="食疗入药月筛选"),
):
    divs = _norm_str_list(division)
    subs = _norm_str_list(subclass)
    ords = _norm_str_list(taxonomic_order)
    fams = _norm_str_list(family)
    gens = _norm_str_list(genus)
    where_sql, params = _build_where(q, divs, subs, ords, fams, gens, harvest_month, food_therapy_month)
    count_sql = f"SELECT COUNT(*) AS c FROM plant_classification_import WHERE {where_sql}"
    list_sql = f"""
        SELECT * FROM plant_classification_import
        WHERE {where_sql}
        ORDER BY id ASC
        LIMIT %s OFFSET %s
    """
    offset = (page - 1) * page_size
    with conn.cursor() as cur:
        cur.execute(count_sql, params)
        total = int((cur.fetchone() or {"c": 0})["c"])
    with conn.cursor() as cur:
        cur.execute(list_sql, [*params, page_size, offset])
        rows = cur.fetchall()
    return PlantListOut(
        items=[plant_row_to_out(r) for r in rows],
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/semantic-search", response_model=PlantListOut)
def semantic_search_plants(
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(get_current_user),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    q: str = Query(..., min_length=1, description="语义搜索的查询词"),
    division: list[str] = Query(default=[], description="门，可多选"),
    subclass: list[str] = Query(default=[], description="亚纲，可多选"),
    taxonomic_order: list[str] = Query(default=[], alias="torder", description="目，可多选"),
    family: list[str] = Query(default=[], description="科，可多选"),
    genus: list[str] = Query(default=[], description="属，可多选"),
    harvest_month: Optional[int] = Query(None, description="最佳采收月筛选"),
    food_therapy_month: Optional[int] = Query(None, description="食疗入药月筛选"),
):
    divs = _norm_str_list(division)
    subs = _norm_str_list(subclass)
    ords = _norm_str_list(taxonomic_order)
    fams = _norm_str_list(family)
    gens = _norm_str_list(genus)
    
    has_filter = bool(divs or subs or ords or fams or gens or harvest_month is not None or food_therapy_month is not None)
    
    allowed_ids = None
    if has_filter:
        where_sql, params = _build_where(None, divs, subs, ords, fams, gens, harvest_month, food_therapy_month)
        ids_sql = f"SELECT id FROM plant_classification_import WHERE {where_sql}"
        with conn.cursor() as cur:
            cur.execute(ids_sql, params)
            rows = cur.fetchall() or []
        allowed_ids = [int(r["id"]) for r in rows]
        if not allowed_ids:
            return PlantListOut(
                items=[],
                total=0,
                page=page,
                page_size=page_size
            )
            
    similarities = query_semantic_similarity(q, allowed_ids=allowed_ids)
    total = len(similarities)
    
    offset = (page - 1) * page_size
    sliced_similarities = similarities[offset : offset + page_size]
    if not sliced_similarities:
        return PlantListOut(
            items=[],
            total=total,
            page=page,
            page_size=page_size
        )
        
    sliced_ids = [item[0] for item in sliced_similarities]
    ph = ",".join(["%s"] * len(sliced_ids))
    list_sql = f"""
        SELECT * FROM plant_classification_import
        WHERE id IN ({ph})
        ORDER BY FIELD(id, {ph})
    """
    with conn.cursor() as cur:
        cur.execute(list_sql, [*sliced_ids, *sliced_ids])
        rows = cur.fetchall() or []
        
    return PlantListOut(
        items=[plant_row_to_out(r) for r in rows],
        total=total,
        page=page,
        page_size=page_size
    )


_BATCH_EXPORT_MAX = 500


@router.post("/export-batch")
def export_plants_batch(
    body: PlantExportBatchIn,
    conn: pymysql.connections.Connection = Depends(get_db),
    user: PlantAdminUser = Depends(get_current_user),
):
    ids = sorted(set(body.ids))
    if len(ids) > _BATCH_EXPORT_MAX:
        raise HTTPException(400, f"一次最多导出 {_BATCH_EXPORT_MAX} 条")
    ph = ",".join(["%s"] * len(ids))
    with conn.cursor() as cur:
        cur.execute(
            f"SELECT * FROM plant_classification_import WHERE id IN ({ph}) ORDER BY id ASC",
            ids,
        )
        rows = cur.fetchall()
    found = {int(r["id"]) for r in rows}
    missing = [i for i in ids if i not in found]
    if missing:
        raise HTTPException(404, f"记录不存在: {missing[:20]}{'…' if len(missing) > 20 else ''}")
    
    _enrich_rows_with_nested_features(rows, conn)
    
    fmt = body.file_format
    log_fmt = "txt" if fmt == "txt" else "xlsx"
    if fmt == "txt":
        payload = plants_export_plain_text(rows).encode("utf-8")
        media_type = "text/plain; charset=utf-8"
        ext = "txt"
    else:
        payload = plants_export_xlsx_bytes(rows)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    with conn.cursor() as cur:
        for r in rows:
            cur.execute(
                """
                INSERT INTO plant_export_logs (plant_id, plant_name, user_id, username, export_format)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (r["id"], r.get("vernacular_name"), user.id, user.username, log_fmt),
            )
    conn.commit()
    filename = f"plants_export_{len(rows)}_{ids[0]}_{ids[-1]}.{ext}"
    return StreamingResponse(
        BytesIO(payload),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/{plant_id}/export")
def export_plant(
    plant_id: int,
    export_fmt: Literal["txt", "xlsx"] = Query("txt", alias="fmt", description="txt 或 xlsx"),
    conn: pymysql.connections.Connection = Depends(get_db),
    user: PlantAdminUser = Depends(get_current_user),
):
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM plant_classification_import WHERE id = %s", (plant_id,))
        p = cur.fetchone()
    if not p:
        raise HTTPException(404, "记录不存在")
    
    _enrich_rows_with_nested_features([p], conn)
    
    log_fmt = "txt" if export_fmt == "txt" else "xlsx"
    if export_fmt == "txt":
        payload = plant_export_plain_text_block(p).encode("utf-8")
        media_type = "text/plain; charset=utf-8"
        ext = "txt"
    else:
        payload = plants_export_xlsx_bytes([p])
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"
    with conn.cursor() as cur:
        cur.execute(
            """
            INSERT INTO plant_export_logs (plant_id, plant_name, user_id, username, export_format)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (p["id"], p.get("vernacular_name"), user.id, user.username, log_fmt),
        )
    conn.commit()
    name = p.get("vernacular_name") or str(p["id"])
    safe = "".join(c if c not in r'\/:*?"<>|' else "_" for c in str(name))[:80]
    filename = f"plant_{safe}_{p['id']}.{ext}"
    return StreamingResponse(
        BytesIO(payload),
        media_type=media_type,
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.post("/{plant_id}/cache-network-image", response_model=PlantOut)
def cache_network_image(
    plant_id: int,
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(require_admin),
):
    """从已保存的「网络图片链接」解析首张图，下载到本站并追加至 image_server_paths。"""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, image_url, image_server_paths FROM plant_classification_import WHERE id = %s",
            (plant_id,),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "记录不存在")
    info = (row.get("image_url") or "").strip()
    if not info:
        raise HTTPException(
            400,
            "未填写网络图片链接，请先在「网络图片链接」中保存后再试",
        )
    existing = _deserialize_paths_from_db(row.get("image_server_paths")) or []
    if len(existing) >= _MAX_MEDIA_PER_PLANT:
        raise HTTPException(400, f"本站图已满（每物种最多 {_MAX_MEDIA_PER_PLANT} 张）")
    img_url, reason = resolve_remote_image_url(info)
    if not img_url:
        raise HTTPException(400, reason)
    try:
        data = download_image_bytes(img_url)
    except urllib.error.HTTPError as e:
        raise HTTPException(502, f"下载图片失败 HTTP {e.code}")
    except OSError as e:
        raise HTTPException(502, f"下载失败: {e}")
    settings = get_settings()
    backend_dir = Path(__file__).resolve().parent.parent.parent
    media_dir = (backend_dir / settings.plant_media_subdir / str(plant_id)).resolve()
    media_dir.mkdir(parents=True, exist_ok=True)
    fname = f"pull_{secrets.token_hex(8)}.jpg"
    dest = media_dir / fname
    dest.write_bytes(data)
    public = f"/api/media/plants/{plant_id}/{fname}"
    merged = existing + [public]
    with conn.cursor() as cur:
        cur.execute(
            "UPDATE plant_classification_import SET image_server_paths=%s WHERE id=%s",
            (_serialize_paths_for_db(merged), plant_id),
        )
    conn.commit()
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM plant_classification_import WHERE id = %s", (plant_id,))
        row = cur.fetchone()
    return plant_row_to_out(row)


@router.post("/{plant_id}/upload-images", response_model=PlantOut)
async def upload_plant_images(
    plant_id: int,
    files: Annotated[list[UploadFile], File(description="可多选 JPG/PNG/WebP/GIF")],
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(get_current_user),
):
    """将多张图片写入服务器目录，并把站内 URL 按顺序追加到 image_server_paths（JSON 数组）。"""
    if not files:
        raise HTTPException(400, "请选择至少一张图片")
    if len(files) > 120:
        raise HTTPException(400, "单次最多上传 120 张")
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, image_server_paths FROM plant_classification_import WHERE id = %s",
            (plant_id,),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "记录不存在")
    existing = _deserialize_paths_from_db(row.get("image_server_paths")) or []
    if len(existing) + len(files) > _MAX_MEDIA_PER_PLANT:
        raise HTTPException(
            400,
            f"超出上限：每物种最多 {_MAX_MEDIA_PER_PLANT} 张，已有 {len(existing)} 张，本次 {len(files)} 张",
        )
    settings = get_settings()
    backend_dir = Path(__file__).resolve().parent.parent.parent
    media_dir = (backend_dir / settings.plant_media_subdir / str(plant_id)).resolve()
    media_dir.mkdir(parents=True, exist_ok=True)
    new_urls: list[str] = []
    for uf in files:
        raw_ct = (uf.content_type or "").split(";")[0].strip().lower()
        ext = _UPLOAD_MIME_TO_EXT.get(raw_ct)
        if not ext:
            raise HTTPException(400, f"不支持的文件类型（仅 JPG/PNG/WebP/GIF）：{uf.filename}")
        blob = await uf.read()
        if len(blob) > _PLANT_UPLOAD_MAX_BYTES:
            raise HTTPException(400, f"单张过大（最大 {_PLANT_UPLOAD_MAX_BYTES // (1024 * 1024)}MB）：{uf.filename}")
        fname = f"i_{secrets.token_hex(8)}{ext}"
        dest = media_dir / fname
        dest.write_bytes(blob)
        new_urls.append(f"/api/media/plants/{plant_id}/{fname}")
    merged = existing + new_urls
    with conn.cursor() as cur:
        cur.execute(
            "UPDATE plant_classification_import SET image_server_paths=%s WHERE id=%s",
            (_serialize_paths_for_db(merged), plant_id),
        )
    conn.commit()
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM plant_classification_import WHERE id = %s", (plant_id,))
        row = cur.fetchone()
    return plant_row_to_out(row)


@router.post("/{plant_id}/delete-server-image", response_model=PlantOut)
def delete_server_image(
    plant_id: int,
    body: PlantDeleteServerImageIn,
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(require_admin),
):
    """仅管理员：从 image_server_paths 中移除给定站内路径并删除磁盘文件（若在托管目录内）。"""
    basename = _basename_under_plant_media(plant_id, body.path)
    if not basename:
        raise HTTPException(400, "path 必须为 /api/media/plants/{id}/下的文件名路径")
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, image_server_paths FROM plant_classification_import WHERE id = %s",
            (plant_id,),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(404, "记录不存在")
    paths = _deserialize_paths_from_db(row.get("image_server_paths")) or []
    target_canon = _strip_url_query(body.path)
    kept: list[str] = []
    removed = False
    for raw in paths:
        if _strip_url_query(raw) == target_canon:
            removed = True
            continue
        kept.append(raw)
    if not removed:
        raise HTTPException(404, "该路径不在 image_server_paths 中")
    settings = get_settings()
    backend_dir = Path(__file__).resolve().parent.parent.parent
    media_dir = (backend_dir / settings.plant_media_subdir / str(plant_id)).resolve()
    disk_path = (media_dir / basename).resolve()
    try:
        disk_path.relative_to(media_dir.resolve())
    except ValueError:
        raise HTTPException(400, "非法路径")
    if disk_path.is_file():
        try:
            disk_path.unlink()
        except OSError as e:
            raise HTTPException(500, f"删除文件失败: {e}") from e
    with conn.cursor() as cur:
        cur.execute(
            "UPDATE plant_classification_import SET image_server_paths=%s WHERE id=%s",
            (_serialize_paths_for_db(kept if kept else None), plant_id),
        )
    conn.commit()
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM plant_classification_import WHERE id = %s", (plant_id,))
        row = cur.fetchone()
    return plant_row_to_out(row)


@router.get("/{plant_id}", response_model=PlantOut)
def get_plant(
    plant_id: int,
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(get_current_user),
):
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM plant_classification_import WHERE id = %s", (plant_id,))
        p = cur.fetchone()
    if not p:
        raise HTTPException(404, "记录不存在")
    with conn.cursor() as cur:
        cur.execute("SELECT * FROM plant_aliases WHERE plant_id = %s", (plant_id,))
        aliases = cur.fetchall() or []
        cur.execute("SELECT habitat_type FROM plant_habitats WHERE plant_id = %s", (plant_id,))
        habitats = [r["habitat_type"] for r in cur.fetchall() or []]
        cur.execute("SELECT * FROM plant_rankings WHERE plant_id = %s", (plant_id,))
        rankings = cur.fetchall() or []
        cur.execute("SELECT * FROM plant_regions WHERE plant_id = %s", (plant_id,))
        regions = cur.fetchall() or []
    return plant_row_to_out(p, aliases=aliases, habitats=habitats, rankings=rankings, regions=regions)


@router.post("", response_model=PlantOut)
def create_plant(
    body: PlantCreate,
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(require_admin),
):
    data = body.model_dump()
    aliases_data = data.pop("aliases", None)
    habitats_data = data.pop("habitats", None)
    rankings_data = data.pop("rankings", None)
    regions_data = data.pop("regions", None)
    
    cols = [c for c in _PLANT_COLS if c in data]
    vals = [_sql_column_value(c, data[c]) for c in cols]
    with conn.cursor() as cur:
        cur.execute("SELECT COALESCE(MAX(id), 0) + 1 AS next_id FROM plant_classification_import")
        new_id = int((cur.fetchone() or {"next_id": 1})["next_id"])
        insert_cols = ["id"] + cols
        insert_vals = [new_id] + vals
        ph = ",".join(["%s"] * len(insert_cols))
        sql = f"INSERT INTO plant_classification_import ({','.join(insert_cols)}) VALUES ({ph})"
        cur.execute(sql, insert_vals)
        
        if aliases_data:
            for item in aliases_data:
                cur.execute(
                    "INSERT INTO plant_aliases (plant_id, alias_type, alias_name, origin_desc) VALUES (%s, %s, %s, %s)",
                    (new_id, item["alias_type"], item["alias_name"], item.get("origin_desc")),
                )
        if habitats_data:
            for h in habitats_data:
                cur.execute(
                    "INSERT INTO plant_habitats (plant_id, habitat_type) VALUES (%s, %s)",
                    (new_id, h),
                )
        if rankings_data:
            for r in rankings_data:
                cur.execute(
                    "INSERT INTO plant_rankings (plant_id, ranking_type, ranking_value, description) VALUES (%s, %s, %s, %s)",
                    (new_id, r["ranking_type"], r.get("ranking_value"), r.get("description")),
                )
        if regions_data:
            for reg in regions_data:
                cur.execute(
                    "INSERT INTO plant_regions (plant_id, region_name, combo_name) VALUES (%s, %s, %s)",
                    (new_id, reg["region_name"], reg.get("combo_name")),
                )
    conn.commit()
    _sync_plant_auto_increment(conn)
    return get_plant(new_id, conn)


@router.put("/{plant_id}", response_model=PlantOut)
def update_plant(
    plant_id: int,
    body: PlantUpdate,
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(require_admin),
):
    with conn.cursor() as cur:
        cur.execute("SELECT id FROM plant_classification_import WHERE id = %s", (plant_id,))
        if not cur.fetchone():
            raise HTTPException(404, "记录不存在")
            
    data = body.model_dump()
    aliases_data = data.pop("aliases", None)
    habitats_data = data.pop("habitats", None)
    rankings_data = data.pop("rankings", None)
    regions_data = data.pop("regions", None)
    
    patch = {k: v for k, v in data.items() if v is not None or k in _PLANT_COLS}
    
    with conn.cursor() as cur:
        if patch:
            sets = ",".join([f"{k}=%s" for k in patch])
            vals = [_sql_column_value(k, v) for k, v in patch.items()] + [plant_id]
            sql = f"UPDATE plant_classification_import SET {sets} WHERE id=%s"
            cur.execute(sql, vals)
            
        if aliases_data is not None:
            cur.execute("DELETE FROM plant_aliases WHERE plant_id = %s", (plant_id,))
            for item in aliases_data:
                cur.execute(
                    "INSERT INTO plant_aliases (plant_id, alias_type, alias_name, origin_desc) VALUES (%s, %s, %s, %s)",
                    (plant_id, item["alias_type"], item["alias_name"], item.get("origin_desc")),
                )
        if habitats_data is not None:
            cur.execute("DELETE FROM plant_habitats WHERE plant_id = %s", (plant_id,))
            for h in habitats_data:
                cur.execute(
                    "INSERT INTO plant_habitats (plant_id, habitat_type) VALUES (%s, %s)",
                    (plant_id, h),
                )
        if rankings_data is not None:
            cur.execute("DELETE FROM plant_rankings WHERE plant_id = %s", (plant_id,))
            for r in rankings_data:
                cur.execute(
                    "INSERT INTO plant_rankings (plant_id, ranking_type, ranking_value, description) VALUES (%s, %s, %s, %s)",
                    (plant_id, r["ranking_type"], r.get("ranking_value"), r.get("description")),
                )
        if regions_data is not None:
            cur.execute("DELETE FROM plant_regions WHERE plant_id = %s", (plant_id,))
            for reg in regions_data:
                cur.execute(
                    "INSERT INTO plant_regions (plant_id, region_name, combo_name) VALUES (%s, %s, %s)",
                    (plant_id, reg["region_name"], reg.get("combo_name")),
                )
    conn.commit()
    return get_plant(plant_id, conn)


@router.delete("/{plant_id}")
def delete_plant(
    plant_id: int,
    conn: pymysql.connections.Connection = Depends(get_db),
    _: PlantAdminUser = Depends(require_admin),
):
    with conn.cursor() as cur:
        cur.execute("DELETE FROM plant_classification_import WHERE id = %s", (plant_id,))
        if cur.rowcount == 0:
            raise HTTPException(404, "记录不存在")
    conn.commit()
    _sync_plant_auto_increment(conn)
    return {"ok": True}


@router.post("/{id}/ai-enrich")
def ai_enrich_plant(
    id: int,
    conn: pymysql.connections.Connection = Depends(get_db),
    current_user: PlantAdminUser = Depends(require_admin),
):
    import urllib.request
    import json

    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, vernacular_name, scientific_name, morphology_text, medicinal_shape, habitat "
            "FROM plant_classification_import WHERE id = %s",
            (id,),
        )
        row = cur.fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="未找到该植物")

    name = row.get("vernacular_name") or ""
    sci_name = row.get("scientific_name") or ""
    morph = row.get("morphology_text") or ""
    med = row.get("medicinal_shape") or ""
    hab = row.get("habitat") or ""

    prompt = f"""你是一个专业的中草药与植物学专家。
根据以下提供的植物信息：
- 中文名称: {name}
- 拉丁名: {sci_name}
- 形态描述: {morph}
- 现有药用性状: {med}
- 生境: {hab}

请帮我深入分析该植物，并生成结构化数据：
1. 最佳采收月份：请列出最适合采收的月份数字，多个月份用逗号分隔，如 "5,6" 或仅一个数字如 "8"。如果没有明确文献，请根据其开花/结果期或通用药用部位（根、茎、叶、花、果）规律进行推断。
2. 最佳采收详细说明：请给出一段详细的中文说明，解释为什么选择这些月份（如开花期、植株成熟度、有效成分含量等），并附带采收和炮制建议（如阴干、烘干等），字数在 100-250 字左右。
3. 适合食疗入药月份：请列出适合作为食疗或药膳入药的月份数字，多个月份用逗号分隔，如 "6,7"。通常与最佳采收月接近或略晚，或为适宜物候调理的季节。
4. 食疗入药详细说明：请给出一段详细的说明，指出食疗或入药的注意事项、鲜品/干品的使用安全风险（如有无毒性、需如何处理如干燥煎煮）、经典古籍记载（如没有特定古籍，可指出同属药材通用逻辑）及服用建议，字数在 150-300 字左右。
5. 药用性状/药用形状说明：如果该植物现有的药用性状描述不完整或无记录，请结合其植物志与药材典籍，推断生成详细的药用形状、性味归经、功效主治说明，字数在 150-300 字左右。
6. 生境场景分类：从 "森林", "草原", "湿地", "荒漠", "海洋" 中选择最适合它的一个或多个分类，如果有其他特殊生境（如 "高山雪线", "农田", "山坡荒地" ），也可以直接生成新分类名称并加入列表中。以字符串数组返回。
7. 特色排行榜推荐：如果该植物具备某些极端或奇特特征（极甜、极苦、国家级珍稀濒危、或生长花期极其漫长/特殊），请为其推荐排行榜分类（ranking_type 必须是: "sweetest"、"bitterest"、"rarity"、"growth_cycle" 之一），并填写指标数值（如 "9.5分", "一级保护"）和详细理由。若均不符合，请返回空数组 []。
8. 道地药材产区关联：若该植物属于某省份的著名道地药材或经典药材名（如 浙江的“浙八味”、河南的“四大怀药”、四川的“川药”、广东的“广药”、吉林的“关药”、云南的“云药”等），请对应生成 region_name (省份) 和 combo_name (经典组合名)。如果不属于上述经典组合，但其在该省份为著名道地特产，可以只填写 region_name 并将 combo_name 设为 null。若均不符合，请返回空数组 []。
9. 分类别名及由来：推荐 1 到 3 个该植物在医药、文学或民间的别称，提供别称类型（必须是："药典标准名", "古书古名", "民间通用俗称", "各地方言名", "药房处方名", "市场商品名", "易混淆错用名" 之一），并详细解释该别名的由来背景或出处（origin_desc）。若无推荐，请返回空数组 []。

请严格以下方的 JSON 格式返回，不要有任何 Markdown 代码块外的文字，也不要包含任何多余信息：
{{
    "harvest_months": "5,6",
    "harvest_months_desc": "最佳采收说明文本...",
    "food_therapy_months": "6,7",
    "food_therapy_months_desc": "食疗入药说明文本...",
    "medicinal_shape": "详细生成的药用形状与功效归经...",
    "habitats": ["森林", "湿地"],
    "rankings": [
        {{
            "ranking_type": "sweetest",
            "ranking_value": "9.5分",
            "description": "含有特异性甜味素，比蔗糖甜数百倍，是著名的天然甜味植物..."
        }}
    ],
    "regions": [
        {{
            "region_name": "浙江",
            "combo_name": "浙八味"
        }}
    ],
    "aliases": [
        {{
            "alias_type": "古书古名",
            "alias_name": "仙草",
            "origin_desc": "在《神农本草经》中被称为仙草，因其具有清热利湿、凉血解毒之奇效..."
        }}
    ]
}}"""

    api_key = "sk-2a32fdc69e89434eafe99a88999380a2"
    payload = {
        "model": "deepseek-v4-flash",
        "messages": [
            {"role": "system", "content": "You are a helpful assistant that outputs only JSON objects."},
            {"role": "user", "content": prompt}
        ],
        "response_format": {"type": "json_object"},
        "temperature": 0.2
    }

    try:
        req = urllib.request.Request(
            "https://api.deepseek.com/v1/chat/completions",
            data=json.dumps(payload).encode("utf-8"),
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {api_key}"
            },
            method="POST"
        )
        with urllib.request.urlopen(req, timeout=30) as response:
            resp_bytes = response.read()
            resp_data = json.loads(resp_bytes.decode("utf-8"))
            content_str = resp_data["choices"][0]["message"]["content"]
            result = json.loads(content_str)
            return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"AI 填充服务发生错误: {str(e)}")
